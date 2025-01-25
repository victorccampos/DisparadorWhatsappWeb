import tkinter as tk
from tkinter import filedialog
import webbrowser
from urllib.parse import quote
import time
import os
from string import Template
import pyautogui
import openpyxl

def selecionar_arquivo(tipo_de_arquivo: str) -> str:
    """
    Função que pede para selecionar o .txt e o .xlsx
    param: tipo_de_arquivo - txt | xlsx
    """
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal

    if tipo_de_arquivo == '.txt':
        filtro = [('Text files', '*.txt')]
    elif tipo_de_arquivo == '.xlsx':
        filtro = [('Excel files', '*.xlsx')]
    else:
        raise ValueError("Tipo de arquivo inválido. Use 'txt' ou 'xlsx'.")

    arquivo = filedialog.askopenfilename(title=f"Selecione um arquivo {tipo_de_arquivo}", filetypes=filtro)

    if arquivo:
        print(f"Arquivo selecionado: {arquivo}")
    else:
        print("Nenhum arquivo selecionado.")
    return arquivo

def get_mensagem_de_texto(texto_path: str) -> Template:
    """Lê o conteúdo de arquivo e o retorna como Template"""
    with open(texto_path, 'r', encoding='utf-8') as arquivo_txt:
        mensagem = arquivo_txt.read().strip()
    return Template(mensagem)
    
def get_mapeamento_colunas(planilha) -> dict[str, int]:
    """ Mapeia os nomes das colunas para seus índices.
    :param planilha: Worksheet do openpyxl
    :return: Dicionário com os nomes das colunas como chaves e os índices como valores
    """
    mapeamento = {}
    for idx, cell in enumerate(planilha[1], start=1):  # Primeira linha como cabeçalho
        if cell.value:  # Ignorar colunas sem nome
            mapeamento[cell.value] = idx
    return mapeamento


if __name__ == '__main__':
    ### planilha, mapeamento_colunas e template_mensagem ###
    
    # Seleciona arquivo de texto, planilha e mapeia índices colunas.
    try:
        texto_path: str = selecionar_arquivo('.txt')
        template_mensagem: Template = get_mensagem_de_texto(texto_path)
        
        workbook_path = selecionar_arquivo('.xlsx')
        workbook = openpyxl.load_workbook(workbook_path)    

        # Mostrar nomes das paginas: 
        nomes_paginas = workbook.sheetnames 
        
        planilha = workbook[nomes_paginas[0]] 
        mapeamento_colunas: dict = get_mapeamento_colunas(planilha)
        INDEX_NOME: int = mapeamento_colunas.get('nome') - 1
        INDEX_TELEFONE: int = mapeamento_colunas.get('telefone') - 1
        # INDEX_EMAIL: int = mapeamento_colunas.get('email') - 1
    except FileNotFoundError as e:
        print(f"Erro: {e}")
    except Exception as e:
        print(f"Erro inesperado: {e}")
    
    ### -------------------------- Envio de mensagens  --------------------- ###
    for linha in planilha.iter_rows(min_row=2):
        NOME: str = linha[INDEX_NOME].value.split(' ')[0]
        TELEFONE: str = str(linha[INDEX_TELEFONE].value)
        
        # Ignora linhas incompletas
        if not NOME or not TELEFONE:  
            continue  
        mensagem = template_mensagem.substitute(nome=NOME, telefone=TELEFONE) 
        try:
            webbrowser.open(f'https://web.whatsapp.com/send?phone={TELEFONE}&text={quote(mensagem)}')
            time.sleep(20)  
            pyautogui.hotkey('enter')  
            time.sleep(5)  # Aguarda o envio da mensagem
            pyautogui.hotkey('ctrl', 'w')  # Fecha a aba do navegador
            time.sleep(3)
        except Exception as e:
            print(f'Erro ao enviar mensagem para {NOME}: {e}')
            # TODO:  Salvar o erro em um arquivo
            with open('erros.csv', 'a', newline='', encoding='utf-8') as arquivo:
                arquivo.write(f'{NOME},{TELEFONE}{os.linesep}')
    print("Envio de mensagens concluído!")
